# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT, relativedelta, pytz
from datetime import datetime, timedelta, date


from datetime import datetime
from tempfile import TemporaryFile
from odoo.exceptions import Warning

import openpyxl as openpyxl
import xlrd
from dateutil.relativedelta import relativedelta
from codecs import encode, decode
import array
import csv
from odoo import api, fields, models, SUPERUSER_ID, _
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT, io
from odoo.tools.float_utils import float_is_zero, float_compare
from odoo.exceptions import UserError, AccessError
from odoo.tools.misc import formatLang
from odoo.addons.base.res.res_partner import WARNING_MESSAGE, WARNING_HELP
# from odoo.addons import decimal_precision as dp
import base64


class PurchaseImport_So(models.Model):
    _name = 'purchase.import.so'
    _description = 'Purchase import to SO'
    # _order = 'order_id, sequence, id'
    # _inherit = 'product.product'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    file_name = fields.Char('Name File',required=True)
    upload_file = fields.Binary('File Upload',required=True)

    @api.multi
    def read_file(self,options):

        # upload_file = self.upload_file.decode('base64')
        upload_file = self.upload_file
        # print(upload_file)
        #work
        # decoded_data = base64.b64decode(self.upload_file)
        # xml_filelike = io.BytesIO(decoded_data)
        # print(xml_filelike)
        # end
        # ---------------------------
        data = base64.b64decode(self.upload_file)
        with open('/tmp/' + self.file_name, 'wb') as file:
            file.write(data)
        xl_workbook = xlrd.open_workbook(file.name)
        sheet_names = xl_workbook.sheet_names()
        xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
        # Number of columns
        num_cols = xl_sheet.ncols
        # header
        headers = []
        # for col_idx in range(0, num_cols):
        #     cell_obj = xl_sheet.cell(0, col_idx)
        #     headers.append(cell_obj.value)
        #
        #     print(xl_sheet.cell(0, col_idx))
        # import_data = []
        # for row_idx in range(1, xl_sheet.nrows):  # Iterate through rows
        #     row_dict = {}
        #     for col_idx in range(0, num_cols):  # Iterate through columns
        #         cell_obj = xl_sheet.cell(row_idx, col_idx)  # Get cell object by row, col
        #         row_dict[headers[col_idx]] = cell_obj.value
        #     import_data.append(row_dict)
        #
        # i = 0
        # for row in import_data:
        #     i += 1
        #     print(str(i)+str(row))
        # --------------------------------------------------
        ##################################################################################################

        Header = ''
        PO_Number = ''
        Order_Date = ''
        Delivery_Date = ''
        TR_code = ''
        Blank_1 = ''
        Delivery_Address = ''
        Department_cus = ''
        Blank_2 = ''
        credit = ''
        Remark = ''

        sale_order_idz              = self.env['sale.order']
        sale_order_line_idz         = self.env['sale.order.line']
        # sale_order_type_idz       = self.env['sale.order.type']

        res_partner_idz             = self.env['res.partner']
        product_template_idz        = self.env['product.product']

        account_payment_term_idz    = self.env['account.payment.term']
        # product_pricelist_idz     = self.env['product.pricelist']
        # stock_warehouse_idz       = self.env['stock.warehouse']

        sale_order = ''


        for row_idx in range(0, xl_sheet.nrows):
            # print(row_idx)

            for col_idx in range(0, num_cols):
                cell_obj = xl_sheet.cell(row_idx, col_idx)
                # print(cell_obj.value)
                ##########################################################################################################
                print('##########################################')
                print('=====================================================')
                txt = str(cell_obj.value)
                # for line_txt in txt:
                #     print(line_txt)
                # print(txt)
                # txt1 = cou
                # print(txt)
                if xl_sheet.nrows:
                    idd = row_idx + 1
                    all_num = xl_sheet.nrows
                    print(idd,'/',all_num)
                # print ()
                ###################### LOG ###############################################
                # msg_body = _("Upload Forecast Non Product  %s Count %s is ") % (0, 0), idd, all_num
                # # self.message_post(body=msg_body, type="notification", subtype="mt_comment")
                # self.message_post(body=msg_body)
                # print(msg_body)
                ##########################################################################
                print('=====================================================')
                # if not txt[0] == 'H':
                #     hhh = idd
                #     print('h',hhh)
                if txt[0] == 'H':
                    # hhh = idd
                    print('H บรรทัดที่ :', idd)
                    ################ print(txt[หลังจาก:ถึง])
                    print('Header           |',txt[0:1])             #Header
                    print('PO Number        |',txt[1:11])            #PO Number
                    print('Order Date       |',txt[11:19])           #Order  Date
                    print('Delivery date    |',txt[19:27])           #Delivery date
                    print('TR code          |',txt[27:47])           #TR code
                    print('Blank_1          |',txt[47:48])           #Blank_1
                    print('Delivery Address |',txt[48:61])           #Delivery Address
                    print('Department cus   |',txt[61:77])           #Department cus
                    print('Blank_2          |',txt[77:82])           #Blank
                    print('credit           |',txt[82:85])           #credit payment_term
                    print('Remark           |',txt[154:227])         #Remark
                    print('bigc_code        |',int(txt[360:373]))  # Remark
                    # print('-------------------------------------------------------------')
                    ans = txt[11:19]
                    arr = str(ans)
                    dmy = str(arr[4])+str(arr[5])+'/'+str(arr[6])+str(arr[7])+'/'+str(arr[0])+str(arr[1])+str(arr[2])+str(arr[3])
                    ans2 = txt[19:27]
                    arr2 = str(ans2)
                    dmy2 = str(arr2[4])+str(arr2[5])+'/'+str(arr2[6])+str(arr2[7])+'/'+ str(arr2[0])+str(arr2[1])+str(arr2[2])+str(arr2[3])
                    # order_date = datetime(dmy)
                    # delivery_date = datetime(dmy2)
                    # print ('Order_date',dmy)
                    # print('Delivery_date', dmy2)
                    # print('================================================')
                    ########################################
                    # sale_order = sale_order_idz.create({
                    #     'partner_id': Department_cus.id,
                    #     # 'partner_invoice_id': credit.id,
                    #     'payment_term_id': credit_payment.id,
                    #     'date_order': dmy,
                    #     'requested_date': dmy2,
                    #     # 'partner_shipping_id': txt[48:61],
                    # })
                    bigc_code_res = res_partner_idz.search([('bigc_code', '=', int(txt[360:373]))])
                    bigc_code_res2 = res_partner_idz.search([('bigc_code', '=', int(txt[48:61]))])
                    credit_payment = account_payment_term_idz.search([('name', '=',txt[82:85])])

                    client_order_ref_log = txt[1:11]
                    note_log = txt[154:227]

                    # origin_id = txt[1:11]
                    # client_order_ref_id = txt[61:77]
                    # testz = sale_order_idz.search([('partner_id', '=', '1112')])
                    # 8859061199992

                    # if Department_cus:
                    sale_order = sale_order_idz.create({
                        'partner_id': bigc_code_res.id,
                        # 'partner_invoice_id': bigc_code_res.id,
                        'partner_shipping_id': bigc_code_res2.id,
                        'payment_term_id': credit_payment.id,
                        'date_order': dmy,
                        'requested_date': dmy2,
                        'client_order_ref': txt[1:11],
                        # 'client_order_ref': credit_payment.id,
                        'note': txt[154:227],
                    })

                    msg_body = _("Upload // <b>Head บรรทัดที่: %s </b>"
                                 "<br/>// Customer: %s "
                                 "<br/>// Delivery Address: %s"
                                 "<br/>// payment_term: %s"
                                 "<br/>// Order Date: %s"
                                 "<br/>// Requested Date: %s"
                                 "<br/>// Customer Reference: %s"
                                 "<br/><b>###########################################################################################</b>") % (idd,bigc_code_res.name,bigc_code_res2.name,credit_payment.name,dmy,dmy2,client_order_ref_log)
                    # ()(dmy2)(client_order_ref_log)
                    self.message_post(body=msg_body)

                    # if not bigc_code_res:
                    #     raise Warning(_('bigc_code_res : ไม่ตรง หรือ ไม่มีอยู่ในระบบ'))
                    # if not bigc_code_res2:
                    #     raise Warning(_('bigc_code_res2 : ไม่ตรง หรือ ไม่มีอยู่ในระบบ'))
                if txt[0] == 'T':
                    print('Detail           |', txt[0:1])  # Header
                    print('PO Number        |', txt[1:11])  # PO Number
                    print('Order            |', txt[11:14])  # Order  Date
                    print('Product ID       |', txt[14:27])  # Delivery date
                    print('Blank            |', txt[27:28])  # TR code
                    print('Product CODE BigC|', txt[28:41])  # Blank_1
                    print('Blank            |', txt[41:42])  # Delivery Address
                    print('Descript product |', txt[42:97])  # Department cus
                    print('Product QTY      |', float(txt[97:107]))  # Blank
                    print('UOM              |', txt[107:117])  # credit
                    print('Unit Price       |', float(txt[117:130]))  # Remark
                    print('Production QTY   |', float(txt[130:142]))  # Remark
                    print('ราคา  หน่วย       |', float(txt[339:350]))
                    print('จำนวน ต่อหน่วย      |', float(txt[351:363]))
                    # print('-------------------------------------------------------------')
                    # print ('TEST',str(txt[42:97]))
                    # print('========================================================')

                    if sale_order:
                        sale_order_line = self.env['sale.order.line']
                        product_idx = product_template_idz.search([('barcode', '=', int(txt[14:27]))])
                        descri = txt[28:97]
                        line_descri = str(descri)
                        if not product_idx:
                            raise UserError(_('Please check product %s') % txt[14:27])
                        # val = {
                        #     'order_id': sale_order.id,                 #ค้นหาตาม H
                        #     'product_id': product_idx.id,
                        #     # 'product_id': int(txt[14:27]),
                        #     'price_unit': float(txt[351:363]),
                        #     'product_uom_qty': float(txt[339:350]),
                        #     'name': line_descri,
                        # }
                        # print (val)
                        # print ('-xxxx')

                        sale_order_line.create({
                            'order_id': sale_order.id,                 #ค้นหาตาม H
                            'product_id': product_idx.id,
                            # 'product_id': int(txt[14:27]),
                            'price_unit': float(txt[351:363]),
                            'product_uom_qty': float(txt[339:350]),
                            'name': line_descri,
                        })
                        bigc_codeee = txt[28:41]
                        msg_body = _("Upload // Line บรรทัดที่ %s //"
                                     "<br/>//Product ID: %s"
                                     "<br/>//Description: %s"
                                     "<br/>//BigC-Code: %s") % (idd,product_idx.name,line_descri,bigc_codeee)
                        self.message_post(body=msg_body)
                print('##########################################')


##########################################################################################################





