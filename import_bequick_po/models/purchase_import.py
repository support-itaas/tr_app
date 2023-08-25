# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from odoo import api, fields, models, _
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT, relativedelta, pytz
from datetime import datetime, timedelta, date


from datetime import datetime
from tempfile import TemporaryFile
from odoo.exceptions import Warning

import openpyxl as openpyxl
import xlrd
from dateutil.relativedelta import relativedelta
from codecs import encode, decode
import array
import csv
from odoo import api, fields, models, SUPERUSER_ID, _
from odoo.tools import DEFAULT_SERVER_DATETIME_FORMAT, io
from odoo.tools.float_utils import float_is_zero, float_compare
from odoo.exceptions import UserError, AccessError
from odoo.tools.misc import formatLang
from odoo.addons.base.res.res_partner import WARNING_MESSAGE, WARNING_HELP
# from odoo.addons import decimal_precision as dp
import base64


class PurchaseImport_So(models.Model):
    _name = 'purchase.import.so.bq'
    _description = 'Purchase import to SO'
    # _order = 'order_id, sequence, id'
    # _inherit = 'product.product'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    file_name = fields.Char('Name File',required=True)
    upload_file = fields.Binary('File Upload',required=True)

    @api.multi
    def read_file(self,options):



        upload_file = self.upload_file
        data = base64.b64decode(self.upload_file)
        with open('/tmp/' + self.file_name, 'wb') as file:
            file.write(data)
        xl_workbook = xlrd.open_workbook(file.name)
        sheet_names = xl_workbook.sheet_names()
        xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
        # Number of columns
        num_cols = xl_sheet.ncols

        # header
        headers = []
        Header = ''
        PO_Number = ''
        Order_Date = ''
        Delivery_Date = ''
        TR_code = ''
        Blank_1 = ''
        Delivery_Address = ''
        Department_cus = ''
        Blank_2 = ''
        credit = ''
        Remark = ''
        sale_order_idz              = self.env['sale.order']
        sale_order_line_idz         = self.env['sale.order.line']
        # sale_order_type_idz       = self.env['sale.order.type']

        res_partner_idz             = self.env['res.partner']
        product_template_idz        = self.env['product.product']

        account_payment_term_idz    = self.env['account.payment.term']
        # product_pricelist_idz     = self.env['product.pricelist']
        # stock_warehouse_idz       = self.env['stock.warehouse']

        sale_order = ''
        sale_order_line_id = ''

        print(xl_sheet.cell(1, 3))
        print(xl_sheet.nrows)
        print(num_cols)
        count= 0
        print('11111111111111111111111111')
        sale_ids = []
        productzzz = ''
        sso = ''
        roadsss = ''
        product_ref = ''
        qty_productzz = 0
        count_done = 0
        count_fail = 0

        productzzz = xl_sheet.cell(4, 1)
        product_za = str(productzzz.value)
        print(product_za)
        print('TEST#@#')
        product_za = self.env['product.product'].search([('default_code', '=', product_za)])
        print('1212121333333333333')
        for col_idx in range(3, num_cols):
            for row_idx in range(1, xl_sheet.nrows):
                # print(row_idx)
                # print(col_idx)
                # print('____________________')
                cell_obj = xl_sheet.cell(row_idx, col_idx)
                txt = str(cell_obj.value)

                if col_idx >=3  and row_idx ==1:
                    sso = self.env['res.partner'].search([('bq_code', '=', txt)])
                    print(sso)
                if col_idx >= 3 and row_idx ==2:
                    roadsss = txt
                if col_idx >=3  and row_idx ==3 :
                    product_ref = txt
                if col_idx >= 3 and row_idx == 4:
                    qty_productzz = txt




            if sso:
                count_done += 1

                vals = {
                    'partner_id': sso.id,
                    'client_order_ref': product_ref,
                    'origin': roadsss,
                    'payment_term_id': 3,

                }
                print(vals)

                order_id = sale_order_idz.create(vals)

                valsa = {
                    'product_id':product_za.id,
                    'name': product_za.name,
                    'product_uom_qty': qty_productzz,
                    'product_uom' : product_za.uom_id.id,
                    'price_unit' : product_za.list_price,
                    'order_id' : order_id.id,
                    'discount' : 50,
                }

                sale_order_line_idz.create(valsa)
            else:
                count_fail += 1
                print(roadsss)

        print(count_done)
        print(count_fail)

















